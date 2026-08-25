#!/usr/bin/env python3
"""Generate content.xlsx for the academic homepage (English version).

Edit the data structures below, then run:
    python generate_content.py
to regenerate content.xlsx. All content the website shows lives in this file.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

OUT = "content.xlsx"

# ---- Profile (key -> value) ---------------------------------------------
profile = [
    ("name", "[Your Name]"),
    ("title", "Ph.D. Candidate in Linguistics"),
    ("affiliation", "[Your University / Department]"),
    ("bio",
     "I am a linguistics researcher working on word meaning, polysemy, and "
     "semantic representation. My current work explores how minimalist semantics "
     "accounts for the representation of polysemous words. Replace this paragraph "
     "with a short, first-person introduction (2-4 sentences)."),
    ("photo", "photo.svg"),
    ("email", "you@example.edu"),
    ("googleScholar", ""),
    ("github", ""),
    ("linkedin", ""),
    ("cv", "CV.pdf"),
]

# ---- Publications (table) -----------------------------------------------
# Columns: Year | Title | Authors | Venue | Link
publications_header = ["Year", "Title", "Authors", "Venue", "Link"]
publications = [
    ["2026", "Semantic Minimalism and the Representation of Polysemy",
     "[Your Name], [Co-author]", "Journal of Semantics", ""],
    ["2025", "A Minimalist Account of Word Meaning",
     "[Your Name]", "Proceedings of the 47th Annual Conference of the Cognitive Science Society", ""],
    ["2024", "Polysemy in Construction Grammar: Evidence from Corpus Data",
     "[Your Name], [Co-author]", "Language", ""],
]

# ---- Research (table) ---------------------------------------------------
# Columns: Title | Description | Image
research_header = ["Title", "Description", "Image"]
research = [
    ["Semantic Minimalism",
     "Investigating how a minimalist semantic framework can capture the core "
     "meaning of polysemous words without over-generation.", ""],
    ["Polysemy & Representation",
     "Building computational models of how multiple related senses are stored "
     "and retrieved in the mental lexicon.", ""],
]

# ---- News (table) ------------------------------------------------------
# Columns: Date | Text
news_header = ["Date", "Text"]
news = [
    ["2026-08", "Launched my academic homepage."],
    ["2026-05", "Paper accepted to the Journal of Semantics."],
    ["2025-11", "Presented at the Annual Meeting of the Linguistics Society."],
]

# ---- Any extra sheet becomes its own section automatically. ------------
# This is just a DEMO. Add/remove/rename sheets freely in Excel; each non-
# Profile sheet shows up as a tab + section with no code changes.
# Generic columns understood: Title, Name, Date, Year, Description, Text.
teaching_header = ["Course", "Term", "Role", "Description"]
teaching = [
    ["Introduction to Linguistics", "Fall 2025", "Teaching Assistant",
     "Led discussion sections and graded assignments for 120 undergraduates."],
    ["Semantics Seminar", "Spring 2025", "Guest Lecturer",
     "Delivered two lectures on minimalist semantics and polysemy."],
]


def style_sheet(ws, header_row=None):
    if header_row:
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2C3E50")
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = f"A{header_row + 1}"


wb = openpyxl.Workbook()

# Profile sheet (key-value)
ws = wb.active
ws.title = "Profile"
ws.append(["Key", "Value"])
for k, v in profile:
    ws.append([k, v])
style_sheet(ws, 1)

# Publications sheet
ws = wb.create_sheet("Publications")
ws.append(publications_header)
for row in publications:
    ws.append(row)
style_sheet(ws, 1)

# Research sheet
ws = wb.create_sheet("Research")
ws.append(research_header)
for row in research:
    ws.append(row)
style_sheet(ws, 1)

# News sheet
ws = wb.create_sheet("News")
ws.append(news_header)
for row in news:
    ws.append(row)
style_sheet(ws, 1)

# Teaching sheet (demo extra section -> auto-rendered)
ws = wb.create_sheet("Teaching")
ws.append(teaching_header)
for row in teaching:
    ws.append(row)
style_sheet(ws, 1)

wb.save(OUT)
print(f"Wrote {OUT} with sheets: {wb.sheetnames}")
